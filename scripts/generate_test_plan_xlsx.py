"""
Genera un libro de Excel con el plan de pruebas para validar el funcionamiento
del sistema EUROPCAR V1.

Salida: Contexto/Plan_de_Pruebas_EUROPCAR_V1.xlsx

El libro incluye tres hojas:
  1. Resumen          -> portada + leyenda de columnas y prioridades.
  2. Plan de pruebas  -> casos de prueba detallados (funcional, seguridad,
                          negativa, integracion y validacion de negocio).
  3. Matriz cobertura -> conteo de casos por modulo y por tipo de prueba.
"""

from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT = REPO_ROOT / "Contexto" / "Plan_de_Pruebas_EUROPCAR_V1.xlsx"


# ---------------------------------------------------------------------------
# Estilo
# ---------------------------------------------------------------------------

PRIMARY = "0F4C81"
PRIMARY_LIGHT = "DCE7F2"
ALT_ROW = "F4F8FC"
HEADER_TEXT = "FFFFFF"
DARK_TEXT = "1B2631"

THIN = Side(style="thin", color="BFC9D1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header_font() -> Font:
    return Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)


def cell_font() -> Font:
    return Font(name="Calibri", size=10, color=DARK_TEXT)


def title_font() -> Font:
    return Font(name="Calibri", size=20, bold=True, color=PRIMARY)


def subtitle_font() -> Font:
    return Font(name="Calibri", size=12, italic=True, color="4A5568")


# ---------------------------------------------------------------------------
# Definicion de casos de prueba
# ---------------------------------------------------------------------------

COLUMNS = [
    ("id", "ID", 8),
    ("modulo", "Modulo", 18),
    ("caso_uso", "Caso de uso", 12),
    ("escenario", "Escenario", 38),
    ("tipo", "Tipo", 14),
    ("prioridad", "Prioridad", 11),
    ("precondiciones", "Precondiciones", 32),
    ("pasos", "Pasos", 50),
    ("datos", "Datos de entrada", 28),
    ("resultado_esperado", "Resultado esperado", 42),
    ("estado", "Estado", 11),
    ("severidad", "Severidad si falla", 14),
    ("responsable", "Responsable", 14),
    ("fecha_ejecucion", "Fecha ejecucion", 14),
    ("comentarios", "Comentarios / evidencia", 32),
]


# Helper: arma un dict con valores por defecto vacios
def case(**overrides) -> dict:
    base = {key: "" for key, _, _ in COLUMNS}
    base["estado"] = "Pendiente"
    base["responsable"] = ""
    base.update(overrides)
    return base


TESTS: list[dict] = []


# 1) Autenticacion / sesion ----------------------------------------------------
TESTS += [
    case(
        id="TC-AUTH-001", modulo="Autenticacion", caso_uso="CU-01",
        escenario="Login exitoso de administrador",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Usuario admin.dev existe, activo, con rol ADMIN.",
        pasos="1. Abrir /login.\n2. Ingresar admin.dev / 12345.\n3. Enviar formulario.",
        datos="username=admin.dev, password=12345",
        resultado_esperado="200 OK con JWT, redirige a /dashboard y la sesion queda persistida en localStorage.",
        severidad="Critica",
    ),
    case(
        id="TC-AUTH-002", modulo="Autenticacion", caso_uso="CU-01",
        escenario="Login con credenciales invalidas",
        tipo="Negativa", prioridad="Alta",
        precondiciones="Usuario existe.",
        pasos="1. POST /api/v1/Auth/login con password incorrecta.",
        datos="username=admin.dev, password=Erronea1",
        resultado_esperado="401 Unauthorized, sin token; mensaje generico (sin revelar detalles); intento auditado en aud_intentos_login.",
        severidad="Alta",
    ),
    case(
        id="TC-AUTH-003", modulo="Autenticacion", caso_uso="CU-01",
        escenario="Bloqueo por intentos fallidos",
        tipo="Seguridad", prioridad="Alta",
        precondiciones="Usuario activo.",
        pasos="1. Realizar 5 intentos fallidos seguidos.\n2. Intentar login nuevamente con password correcta.",
        datos="username=cliente.web, password incorrecta x5",
        resultado_esperado="La cuenta queda bloqueada hasta BloqueadoHastaUtc; el login posterior responde 401 con motivo de bloqueo.",
        severidad="Critica",
    ),
    case(
        id="TC-AUTH-004", modulo="Autenticacion", caso_uso="CU-02",
        escenario="Registro de cliente nuevo",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Cedula no registrada previamente.",
        pasos="1. Abrir /registro.\n2. Completar formulario.\n3. Enviar.",
        datos="cedula=1700000001, nombre=Juan, password=Test1234",
        resultado_esperado="200 OK, cliente y usuario CLIENTE_WEB creados; el frontend redirige al portal del cliente.",
        severidad="Alta",
    ),
    case(
        id="TC-AUTH-005", modulo="Autenticacion", caso_uso="CU-02",
        escenario="Registro con cedula duplicada",
        tipo="Negativa", prioridad="Media",
        precondiciones="Existe cliente con la cedula.",
        pasos="1. GET /api/v1/Auth/cedula-exists?cedula=...\n2. POST /api/v1/Auth/register con misma cedula.",
        datos="cedula existente",
        resultado_esperado="cedula-exists devuelve exists=true; el POST register retorna 409/422 con mensaje claro.",
        severidad="Media",
    ),
    case(
        id="TC-AUTH-006", modulo="Autenticacion", caso_uso="CU-01",
        escenario="JWT expirado",
        tipo="Seguridad", prioridad="Alta",
        precondiciones="Token con expiracion vencida (esperar > ExpirationMinutes).",
        pasos="1. Esperar a que expire el token.\n2. Llamar a un endpoint protegido.",
        datos="Authorization: Bearer <expirado>",
        resultado_esperado="401 Unauthorized; el frontend limpia sesion y redirige a /login conservando 'from'.",
        severidad="Alta",
    ),
    case(
        id="TC-AUTH-007", modulo="Autenticacion", caso_uso="CU-01",
        escenario="Acceso a rutas protegidas sin token",
        tipo="Seguridad", prioridad="Alta",
        precondiciones="Sesion cerrada.",
        pasos="1. Navegar a /dashboard sin estar autenticado.",
        datos="-",
        resultado_esperado="ProtectedRoute redirige a /login; el endpoint backend responde 401 si se invoca directo.",
        severidad="Alta",
    ),
    case(
        id="TC-AUTH-008", modulo="Autenticacion", caso_uso="CU-11",
        escenario="Usuario CLIENTE_WEB intenta acceder a panel admin",
        tipo="Seguridad", prioridad="Alta",
        precondiciones="Sesion activa con rol CLIENTE_WEB.",
        pasos="1. Navegar a /vehiculos.\n2. Llamar GET /api/v1/admin/vehiculos.",
        datos="JWT con rol CLIENTE_WEB",
        resultado_esperado="ProtectedRoute redirige a /mi-cuenta; el backend retorna 403 para el endpoint admin.",
        severidad="Alta",
    ),
]

# 2) Catalogo / Busqueda -------------------------------------------------------
TESTS += [
    case(
        id="TC-CAT-001", modulo="Catalogo", caso_uso="CU-03",
        escenario="Buscar vehiculos disponibles por fechas y localizacion",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Existen vehiculos activos en la sucursal.",
        pasos="1. GET /api/v1/vehiculos?idLocalizacion=1&fechaRecogida=...&fechaDevolucion=...",
        datos="fechas validas, lead time cumplido",
        resultado_esperado="200 OK con lista paginada de vehiculos disponibles, sin incluir reservados ni en taller.",
        severidad="Alta",
    ),
    case(
        id="TC-CAT-002", modulo="Catalogo", caso_uso="CU-03",
        escenario="Detalle de vehiculo",
        tipo="Funcional", prioridad="Media",
        precondiciones="Vehiculo existe y esta activo.",
        pasos="1. GET /api/v1/vehiculos/{vehiculoId}",
        datos="vehiculoId=veh-001",
        resultado_esperado="200 OK con marca, categoria, capacidades, precio_dia, foto, sucursal.",
        severidad="Media",
    ),
    case(
        id="TC-CAT-003", modulo="Catalogo", caso_uso="CU-03",
        escenario="Disponibilidad en tiempo real",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Vehiculo existe.",
        pasos="1. GET /api/v1/vehiculos/{vehiculoId}/disponibilidad",
        datos="fechas validas",
        resultado_esperado="Disponible=true cuando no hay solapamiento; false cuando hay reservas/contratos en el rango.",
        severidad="Alta",
    ),
    case(
        id="TC-CAT-004", modulo="Catalogo", caso_uso="CU-03",
        escenario="Listar localizaciones publicas",
        tipo="Funcional", prioridad="Media",
        precondiciones="Existen localizaciones activas.",
        pasos="1. GET /api/v1/localizaciones",
        datos="-",
        resultado_esperado="200 OK con sucursales activas, ordenadas y con datos de contacto.",
        severidad="Media",
    ),
    case(
        id="TC-CAT-005", modulo="Catalogo", caso_uso="CU-03",
        escenario="Listar categorias y extras",
        tipo="Funcional", prioridad="Baja",
        precondiciones="Catalogos cargados.",
        pasos="1. GET /api/v1/categorias.\n2. GET /api/v1/extras.",
        datos="-",
        resultado_esperado="200 OK con categorias y extras (con valor fijo).",
        severidad="Media",
    ),
]

# 3) Reservas ------------------------------------------------------------------
TESTS += [
    case(
        id="TC-RES-001", modulo="Reservas", caso_uso="CU-03",
        escenario="Crear reserva como cliente invitado",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Vehiculo disponible en el rango.",
        pasos="1. POST /api/v1/admin/reservas/guest-client.\n2. POST /api/v1/admin/reservas con idCliente devuelto.",
        datos="cedula, nombre, fechas, idVehiculo, idLocalizaciones",
        resultado_esperado="Reserva creada en estado PENDIENTE con codigo unico; calculo correcto de subtotal, impuestos, extras y total.",
        severidad="Critica",
    ),
    case(
        id="TC-RES-002", modulo="Reservas", caso_uso="CU-03",
        escenario="Reserva con fechas invertidas",
        tipo="Negativa", prioridad="Alta",
        precondiciones="Vehiculo activo.",
        pasos="1. POST /api/v1/admin/reservas con fechaDevolucion < fechaRecogida.",
        datos="fechas invertidas",
        resultado_esperado="422 con error 'Las fechas son invalidas' y no se persiste reserva.",
        severidad="Alta",
    ),
    case(
        id="TC-RES-003", modulo="Reservas", caso_uso="CU-03",
        escenario="Prevencion de overbooking",
        tipo="Integracion", prioridad="Alta",
        precondiciones="Una reserva CONFIRMADA cubre el rango T1-T2.",
        pasos="1. Intentar crear nueva reserva del mismo vehiculo en T1-T2.",
        datos="idVehiculo en uso",
        resultado_esperado="409 Conflict con motivo de solapamiento; no se crea segunda reserva.",
        severidad="Critica",
    ),
    case(
        id="TC-RES-004", modulo="Reservas", caso_uso="CU-04",
        escenario="Confirmar reserva pendiente",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Reserva en PENDIENTE.",
        pasos="1. PUT /api/v1/admin/reservas/{id}/confirmar con monto y referencia.",
        datos="id valido, monto=total",
        resultado_esperado="Reserva pasa a CONFIRMADA; se generan Pago (PAGADO) y Factura (EMITIDA) en una transaccion.",
        severidad="Critica",
    ),
    case(
        id="TC-RES-005", modulo="Reservas", caso_uso="CU-05",
        escenario="Cancelar reserva propia con fecha futura",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Reserva del cliente con fecha de retiro futura.",
        pasos="1. PUT /api/v1/admin/reservas/{id}/cancelar con motivo.",
        datos="motivo=Cambio de planes",
        resultado_esperado="Reserva pasa a CANCELADA; pagos y facturas asociados se anulan; queda registro de auditoria.",
        severidad="Alta",
    ),
    case(
        id="TC-RES-006", modulo="Reservas", caso_uso="CU-05",
        escenario="Cliente intenta cancelar reserva de otro cliente",
        tipo="Seguridad", prioridad="Alta",
        precondiciones="Sesion CLIENTE_WEB sobre reserva ajena.",
        pasos="1. PUT /api/v1/admin/reservas/{id}/cancelar.",
        datos="reserva ajena",
        resultado_esperado="403 Forbidden con mensaje 'No puedes cancelar reservas que no te pertenecen'.",
        severidad="Alta",
    ),
    case(
        id="TC-RES-007", modulo="Reservas", caso_uso="CU-05",
        escenario="Cliente intenta cancelar reserva pasada",
        tipo="Negativa", prioridad="Media",
        precondiciones="Reserva del cliente con fecha de recogida pasada.",
        pasos="1. PUT /api/v1/admin/reservas/{id}/cancelar.",
        datos="reserva pasada",
        resultado_esperado="403/409 con mensaje informando que ya no se puede cancelar.",
        severidad="Media",
    ),
    case(
        id="TC-RES-008", modulo="Reservas", caso_uso="CU-12",
        escenario="Cliente consulta solo sus reservas",
        tipo="Seguridad", prioridad="Alta",
        precondiciones="Sesion CLIENTE_WEB.",
        pasos="1. GET /api/v1/admin/reservas/cliente/{idCliente} con id distinto al del token.",
        datos="-",
        resultado_esperado="403 Forbidden; con su mismo id retorna 200 OK con su listado.",
        severidad="Alta",
    ),
]

# 4) Contratos ----------------------------------------------------------------
TESTS += [
    case(
        id="TC-CON-001", modulo="Contratos", caso_uso="CU-06",
        escenario="Apertura de contrato desde reserva confirmada",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Reserva CONFIRMADA del dia.",
        pasos="1. POST /api/v1/Contratos con datos de salida (kilometraje, combustible).",
        datos="idReserva valido",
        resultado_esperado="Contrato ABIERTO con numero unico; vehiculo cambia a estado ALQUILADO.",
        severidad="Alta",
    ),
    case(
        id="TC-CON-002", modulo="Contratos", caso_uso="CU-06",
        escenario="Check-out sin licencia vigente",
        tipo="Negativa", prioridad="Media",
        precondiciones="Conductor con licencia vencida.",
        pasos="1. POST /api/v1/Contratos/checkout.",
        datos="conductor con licencia vencida",
        resultado_esperado="422 con regla de negocio 'Licencia vencida'; el vehiculo no se entrega.",
        severidad="Alta",
    ),
    case(
        id="TC-CON-003", modulo="Contratos", caso_uso="CU-07",
        escenario="Check-in sin recargos",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Contrato ABIERTO; devolucion en horario y limpio.",
        pasos="1. POST /api/v1/Contratos/checkin.",
        datos="km dentro de limite, combustible lleno, limpio=true",
        resultado_esperado="Contrato CERRADO; vehiculo DISPONIBLE; sin pagos adicionales generados.",
        severidad="Alta",
    ),
    case(
        id="TC-CON-004", modulo="Contratos", caso_uso="CU-07",
        escenario="Check-in con cargos por combustible y km extra",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Contrato ABIERTO; vehiculo devuelto sin combustible y con km excedente.",
        pasos="1. POST /api/v1/Contratos/checkin con esos datos.",
        datos="combustible=0, km > limite",
        resultado_esperado="Se generan cargos calculados y un Pago adicional; el contrato pasa a CERRADO.",
        severidad="Alta",
    ),
    case(
        id="TC-CON-005", modulo="Contratos", caso_uso="CU-06",
        escenario="Check-out sobre reserva no confirmada",
        tipo="Negativa", prioridad="Alta",
        precondiciones="Reserva en estado PENDIENTE.",
        pasos="1. POST /api/v1/Contratos/checkout.",
        datos="reserva PENDIENTE",
        resultado_esperado="409/422 indicando que la reserva debe estar CONFIRMADA antes del check-out.",
        severidad="Alta",
    ),
]

# 5) Pagos y facturas ----------------------------------------------------------
TESTS += [
    case(
        id="TC-PAG-001", modulo="Pagos", caso_uso="CU-04",
        escenario="Registrar pago manual",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Reserva o contrato existente.",
        pasos="1. POST /api/v1/Pagos con monto, metodo y referencia.",
        datos="monto=total, metodo=TARJETA",
        resultado_esperado="201 Created con codigo de pago, asociado a la reserva/contrato.",
        severidad="Alta",
    ),
    case(
        id="TC-PAG-002", modulo="Pagos", caso_uso="CU-04",
        escenario="Listar pagos por reserva",
        tipo="Funcional", prioridad="Media",
        precondiciones="Existen pagos asociados.",
        pasos="1. GET /api/v1/Pagos/reserva/{idReserva}.",
        datos="-",
        resultado_esperado="200 OK con todos los pagos (PAGADO/ANULADO) ordenados por fecha.",
        severidad="Media",
    ),
    case(
        id="TC-FAC-001", modulo="Facturas", caso_uso="CU-12",
        escenario="Cliente consulta sus facturas",
        tipo="Funcional", prioridad="Media",
        precondiciones="Cliente con facturas EMITIDAS.",
        pasos="1. GET /api/v1/Facturas/mis-facturas.",
        datos="JWT CLIENTE_WEB",
        resultado_esperado="200 OK con facturas del cliente solicitante; no devuelve facturas de otros clientes.",
        severidad="Media",
    ),
]

# 6) Vehiculos / Flota ---------------------------------------------------------
TESTS += [
    case(
        id="TC-VEH-001", modulo="Vehiculos", caso_uso="CU-08",
        escenario="Crear vehiculo en flota",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Sesion ADMIN/AGENTE_POS.",
        pasos="1. POST /api/v1/admin/vehiculos.",
        datos="placa, marca, categoria, anio, precio_base",
        resultado_esperado="201 con vehiculo creado, codigoInternoVehiculo unico y estado DISPONIBLE.",
        severidad="Alta",
    ),
    case(
        id="TC-VEH-002", modulo="Vehiculos", caso_uso="CU-08",
        escenario="Crear vehiculo con placa duplicada",
        tipo="Negativa", prioridad="Alta",
        precondiciones="Existe un vehiculo con la misma placa.",
        pasos="1. POST /api/v1/admin/vehiculos con placa repetida.",
        datos="placa duplicada",
        resultado_esperado="409/422 con mensaje 'Placa duplicada'; no se persiste.",
        severidad="Alta",
    ),
    case(
        id="TC-VEH-003", modulo="Vehiculos", caso_uso="CU-08",
        escenario="Editar vehiculo y subir imagen",
        tipo="Funcional", prioridad="Media",
        precondiciones="Vehiculo existente.",
        pasos="1. PUT /api/v1/admin/vehiculos/{id} con ImagenReferencialUrl.",
        datos="imagen Cloudinary",
        resultado_esperado="200 OK con imagen actualizada y datos modificados.",
        severidad="Media",
    ),
    case(
        id="TC-VEH-004", modulo="Vehiculos", caso_uso="CU-08",
        escenario="Eliminar vehiculo con reservas activas",
        tipo="Negativa", prioridad="Alta",
        precondiciones="Vehiculo con reservas futuras.",
        pasos="1. DELETE /api/v1/admin/vehiculos/{id}.",
        datos="-",
        resultado_esperado="409 con mensaje de regla de negocio; el vehiculo no se elimina.",
        severidad="Alta",
    ),
    case(
        id="TC-VEH-005", modulo="Vehiculos", caso_uso="CU-08",
        escenario="Eliminacion logica (soft delete) por ADMIN",
        tipo="Funcional", prioridad="Media",
        precondiciones="Vehiculo sin reservas.",
        pasos="1. DELETE /api/v1/admin/vehiculos/{id} con rol ADMIN.",
        datos="-",
        resultado_esperado="200 OK; EsEliminado=true; queda fuera de listados pero conserva historial.",
        severidad="Media",
    ),
]

# 7) Mantenimientos ------------------------------------------------------------
TESTS += [
    case(
        id="TC-MAN-001", modulo="Mantenimientos", caso_uso="CU-09",
        escenario="Enviar vehiculo a taller",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Vehiculo DISPONIBLE.",
        pasos="1. POST /api/v1/Mantenimientos.",
        datos="idVehiculo, tipo=PREVENTIVO",
        resultado_esperado="201 con mantenimiento ABIERTO; el vehiculo cambia a estado TALLER.",
        severidad="Alta",
    ),
    case(
        id="TC-MAN-002", modulo="Mantenimientos", caso_uso="CU-09",
        escenario="Cerrar mantenimiento",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Mantenimiento ABIERTO.",
        pasos="1. PUT /api/v1/Mantenimientos/{id}/cerrar.",
        datos="-",
        resultado_esperado="Mantenimiento CERRADO; vehiculo regresa a DISPONIBLE.",
        severidad="Alta",
    ),
    case(
        id="TC-MAN-003", modulo="Mantenimientos", caso_uso="CU-09",
        escenario="Mantenimiento de vehiculo ya alquilado",
        tipo="Negativa", prioridad="Media",
        precondiciones="Vehiculo en estado ALQUILADO.",
        pasos="1. POST /api/v1/Mantenimientos.",
        datos="-",
        resultado_esperado="409 con mensaje de regla de negocio; no se permite hasta el check-in.",
        severidad="Media",
    ),
]

# 8) Localizaciones / Sucursales -----------------------------------------------
TESTS += [
    case(
        id="TC-LOC-001", modulo="Localizaciones", caso_uso="CU-10",
        escenario="Crear sucursal",
        tipo="Funcional", prioridad="Media",
        precondiciones="Sesion ADMIN.",
        pasos="1. POST /api/v1/admin/localizaciones.",
        datos="codigo, nombre, idCiudad, direccion, horario",
        resultado_esperado="201 con localizacion ACT y codigo unico.",
        severidad="Media",
    ),
    case(
        id="TC-LOC-002", modulo="Localizaciones", caso_uso="CU-10",
        escenario="Inhabilitar sucursal con vehiculos asignados",
        tipo="Negativa", prioridad="Media",
        precondiciones="Sucursal con vehiculos en localizacion_actual.",
        pasos="1. PUT /api/v1/admin/localizaciones/{id}/estado con INACT.",
        datos="-",
        resultado_esperado="409 indicando que primero deben moverse o desasignarse los vehiculos.",
        severidad="Media",
    ),
    case(
        id="TC-LOC-003", modulo="Localizaciones", caso_uso="CU-10",
        escenario="Listar ciudades para selector",
        tipo="Funcional", prioridad="Baja",
        precondiciones="Catalogo de ciudades cargado.",
        pasos="1. GET /api/v1/admin/localizaciones/ciudades.",
        datos="-",
        resultado_esperado="200 OK con ciudades activas.",
        severidad="Baja",
    ),
]

# 9) Clientes ------------------------------------------------------------------
TESTS += [
    case(
        id="TC-CLI-001", modulo="Clientes", caso_uso="CU-08",
        escenario="Crear cliente desde back-office",
        tipo="Funcional", prioridad="Media",
        precondiciones="Sesion ADMIN/AGENTE_POS.",
        pasos="1. POST /api/v1/Clientes con datos completos.",
        datos="cedula, nombres, apellidos, correo, telefono",
        resultado_esperado="201 con cliente y codigo_cliente unico.",
        severidad="Media",
    ),
    case(
        id="TC-CLI-002", modulo="Clientes", caso_uso="CU-08",
        escenario="Validar duplicidad de identificacion",
        tipo="Negativa", prioridad="Media",
        precondiciones="Cliente con misma cedula.",
        pasos="1. POST /api/v1/Clientes con la misma cedula.",
        datos="cedula duplicada",
        resultado_esperado="409/422 con mensaje claro y sin persistir.",
        severidad="Media",
    ),
    case(
        id="TC-CLI-003", modulo="Clientes", caso_uso="CU-08",
        escenario="Editar correo de cliente",
        tipo="Funcional", prioridad="Baja",
        precondiciones="Cliente existe.",
        pasos="1. PUT /api/v1/Clientes/{id}.",
        datos="correo nuevo",
        resultado_esperado="200 OK con correo actualizado.",
        severidad="Baja",
    ),
]

# 10) Usuarios y roles ---------------------------------------------------------
TESTS += [
    case(
        id="TC-USR-001", modulo="Usuarios", caso_uso="CU-11",
        escenario="Crear usuario interno con rol AGENTE_POS",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Sesion ADMIN.",
        pasos="1. POST /api/v1/Usuarios.",
        datos="username, correo, password, roles=[AGENTE_POS]",
        resultado_esperado="201 con userId; cliente interno asociado y rol asignado correctamente.",
        severidad="Alta",
    ),
    case(
        id="TC-USR-002", modulo="Usuarios", caso_uso="CU-11",
        escenario="Username o correo duplicado",
        tipo="Negativa", prioridad="Alta",
        precondiciones="Existe un usuario con ese username/correo.",
        pasos="1. POST /api/v1/Usuarios duplicado.",
        datos="username/correo existentes",
        resultado_esperado="409 Conflict con mensaje correspondiente.",
        severidad="Alta",
    ),
    case(
        id="TC-USR-003", modulo="Usuarios", caso_uso="CU-11",
        escenario="Cambiar estado de usuario",
        tipo="Funcional", prioridad="Media",
        precondiciones="Usuario existe.",
        pasos="1. PUT /api/v1/Usuarios/{id}/estado.",
        datos="estado=INACT",
        resultado_esperado="200 OK; el usuario inhabilitado no puede iniciar sesion.",
        severidad="Alta",
    ),
    case(
        id="TC-USR-004", modulo="Usuarios", caso_uso="CU-11",
        escenario="Actualizar roles de un usuario",
        tipo="Funcional", prioridad="Media",
        precondiciones="Usuario con rol AGENTE_POS.",
        pasos="1. PUT /api/v1/Usuarios/{id}/roles con [ADMIN].",
        datos="-",
        resultado_esperado="Roles actualizados; el siguiente login refleja los nuevos permisos.",
        severidad="Alta",
    ),
]

# 11) Booking publico ----------------------------------------------------------
TESTS += [
    case(
        id="TC-BKG-001", modulo="Booking", caso_uso="CU-13",
        escenario="Booking crea reserva por canal externo",
        tipo="Integracion", prioridad="Alta",
        precondiciones="Vehiculo disponible.",
        pasos="1. POST /api/v1/reservas con payload publico.",
        datos="codigoVehiculo, fechas, cliente",
        resultado_esperado="201 con codigoReserva; reserva persiste con canal=BOOKING.",
        severidad="Alta",
    ),
    case(
        id="TC-BKG-002", modulo="Booking", caso_uso="CU-13",
        escenario="Cancelar reserva externa via PATCH",
        tipo="Integracion", prioridad="Alta",
        precondiciones="Reserva con canal BOOKING.",
        pasos="1. PATCH /api/v1/reservas/{codigo}/cancelar.",
        datos="motivo",
        resultado_esperado="Reserva CANCELADA; pagos y factura asociados se anulan.",
        severidad="Alta",
    ),
    case(
        id="TC-BKG-003", modulo="Booking", caso_uso="CU-13",
        escenario="Obtener factura de reserva externa",
        tipo="Funcional", prioridad="Media",
        precondiciones="Reserva BOOKING con factura emitida.",
        pasos="1. GET /api/v1/reservas/{codigo}/factura.",
        datos="-",
        resultado_esperado="200 OK con datos de factura.",
        severidad="Media",
    ),
]

# 12) Frontend / UX ------------------------------------------------------------
TESTS += [
    case(
        id="TC-FE-001", modulo="Frontend", caso_uso="CU-01",
        escenario="ProtectedRoute por tipo de usuario",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Sesion CLIENTE_WEB activa.",
        pasos="1. Navegar a /dashboard.",
        datos="-",
        resultado_esperado="Redirige a /mi-cuenta automaticamente.",
        severidad="Alta",
    ),
    case(
        id="TC-FE-002", modulo="Frontend", caso_uso="-",
        escenario="Manejo global de error 500",
        tipo="Negativa", prioridad="Media",
        precondiciones="API respondiendo 500 simulado.",
        pasos="1. Realizar una operacion que dispare 500.",
        datos="-",
        resultado_esperado="Toast de error con traceId; log en consola; sin crash de la app gracias a ErrorBoundary.",
        severidad="Media",
    ),
    case(
        id="TC-FE-003", modulo="Frontend", caso_uso="-",
        escenario="Reintento ante 503 transitorio",
        tipo="Resiliencia", prioridad="Media",
        precondiciones="Endpoint GET con respuesta 503 una sola vez.",
        pasos="1. Realizar GET protegido.",
        datos="-",
        resultado_esperado="axiosClient reintenta hasta 2 veces con backoff y resuelve la peticion.",
        severidad="Media",
    ),
    case(
        id="TC-FE-004", modulo="Frontend", caso_uso="CU-01",
        escenario="Logout limpia sesion",
        tipo="Funcional", prioridad="Alta",
        precondiciones="Sesion activa.",
        pasos="1. Click en Cerrar sesion.",
        datos="-",
        resultado_esperado="useAuthStore.logout(); localStorage limpio; redirige a /login.",
        severidad="Alta",
    ),
    case(
        id="TC-FE-005", modulo="Frontend", caso_uso="CU-03",
        escenario="Validacion de formulario de reserva",
        tipo="Funcional", prioridad="Media",
        precondiciones="Pagina /reservar/:id abierta.",
        pasos="1. Enviar el formulario con campos vacios.",
        datos="-",
        resultado_esperado="Errores por campo via Zod + react-hook-form; no se realiza el POST.",
        severidad="Media",
    ),
]

# 13) Salud / observabilidad ---------------------------------------------------
TESTS += [
    case(
        id="TC-OBS-001", modulo="Salud", caso_uso="-",
        escenario="Liveness",
        tipo="Funcional", prioridad="Media",
        precondiciones="API arriba.",
        pasos="1. GET /health/live.",
        datos="-",
        resultado_esperado="200 OK con status=Healthy en menos de 200 ms.",
        severidad="Alta",
    ),
    case(
        id="TC-OBS-002", modulo="Salud", caso_uso="-",
        escenario="Readiness con BD activa",
        tipo="Funcional", prioridad="Alta",
        precondiciones="BD conectada.",
        pasos="1. GET /health/ready.",
        datos="-",
        resultado_esperado="200 OK con check de BD en estado Healthy.",
        severidad="Alta",
    ),
    case(
        id="TC-OBS-003", modulo="Salud", caso_uso="-",
        escenario="Readiness con BD caida",
        tipo="Resiliencia", prioridad="Alta",
        precondiciones="BD inaccesible.",
        pasos="1. GET /health/ready.",
        datos="-",
        resultado_esperado="503 con detalle del check fallido y traceId.",
        severidad="Alta",
    ),
]

# 14) Rendimiento --------------------------------------------------------------
TESTS += [
    case(
        id="TC-PRF-001", modulo="Rendimiento", caso_uso="-",
        escenario="Latencia de busqueda",
        tipo="Rendimiento", prioridad="Media",
        precondiciones="Carga moderada (50 RPS).",
        pasos="1. Ejecutar prueba de carga sobre GET /api/v1/vehiculos.",
        datos="50 RPS, 5 minutos",
        resultado_esperado="p95 < 500 ms; sin errores 5xx.",
        severidad="Alta",
    ),
    case(
        id="TC-PRF-002", modulo="Rendimiento", caso_uso="-",
        escenario="Concurrencia en confirmacion",
        tipo="Rendimiento", prioridad="Alta",
        precondiciones="Reservas pendientes simultaneas.",
        pasos="1. Confirmar 200 reservas concurrentes.",
        datos="-",
        resultado_esperado="Sin deadlocks; pagos y facturas consistentes; throughput >= 200 TPS objetivo.",
        severidad="Alta",
    ),
    case(
        id="TC-PRF-003", modulo="Rendimiento", caso_uso="-",
        escenario="LCP del frontend",
        tipo="Rendimiento", prioridad="Media",
        precondiciones="Build de produccion del frontend.",
        pasos="1. Medir Web Vitals con Lighthouse.",
        datos="-",
        resultado_esperado="LCP < 2 s en conexion 4G simulada.",
        severidad="Media",
    ),
]

# 15) Auditoria y seguridad -----------------------------------------------------
TESTS += [
    case(
        id="TC-AUD-001", modulo="Auditoria", caso_uso="-",
        escenario="Auditoria de cambios en flota",
        tipo="Seguridad", prioridad="Alta",
        precondiciones="Vehiculo existente.",
        pasos="1. PUT /api/v1/admin/vehiculos/{id} con cambio de precio.",
        datos="precio nuevo",
        resultado_esperado="audit.aud_eventos registra el cambio con datos_anteriores y datos_nuevos.",
        severidad="Alta",
    ),
    case(
        id="TC-AUD-002", modulo="Auditoria", caso_uso="-",
        escenario="Auditoria de intentos de login",
        tipo="Seguridad", prioridad="Alta",
        precondiciones="Usuario activo.",
        pasos="1. Intento exitoso y 1 fallido.",
        datos="-",
        resultado_esperado="Cada intento queda registrado con resultado, IP, user-agent y motivo si aplica.",
        severidad="Alta",
    ),
]


# ---------------------------------------------------------------------------
# Construccion del libro
# ---------------------------------------------------------------------------

def build_resumen(ws) -> None:
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80

    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = "Plan de pruebas - Servicio EUROPCAR V1"
    cell.font = title_font()
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:B2")
    cell = ws["A2"]
    cell.value = "Validacion del funcionamiento del sistema (backend + frontend + base de datos)."
    cell.font = subtitle_font()
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    info = [
        ("Modulos cubiertos",
         "Autenticacion, Catalogo, Reservas, Contratos, Pagos, Facturas, Vehiculos, Mantenimientos, "
         "Localizaciones, Clientes, Usuarios, Booking, Frontend, Salud, Rendimiento y Auditoria."),
        ("Tipos de prueba",
         "Funcional, Negativa, Seguridad, Integracion, Resiliencia, Rendimiento."),
        ("Prioridades", "Alta, Media, Baja."),
        ("Severidad si falla", "Critica, Alta, Media, Baja."),
        ("Estado", "Pendiente, En ejecucion, Pasada, Fallida, Bloqueada."),
        ("Convencion de IDs",
         "TC-<MODULO>-<###> (Ej. TC-AUTH-001). Permite trazabilidad con casos de uso CU-XX."),
        ("Total de casos", str(len(TESTS))),
        ("Generado por", "scripts/generate_test_plan_xlsx.py (regenerable)"),
    ]
    for i, (k, v) in enumerate(info, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, color=PRIMARY)
        ws.cell(row=i, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=i, column=2).font = cell_font()
        ws.row_dimensions[i].height = 32

    ws.cell(row=4 + len(info) + 1, column=1,
            value="Como diligenciar la hoja 'Plan de pruebas'").font = Font(bold=True, color=PRIMARY)
    instructions = [
        "1. Asignar un responsable a cada caso.",
        "2. Ejecutar el caso siguiendo los pasos descritos.",
        "3. Marcar el estado (Pendiente, En ejecucion, Pasada, Fallida, Bloqueada).",
        "4. Si falla, registrar severidad, comentarios y referencia a la evidencia (captura, log, trace).",
        "5. Actualizar la fecha de ejecucion en formato YYYY-MM-DD.",
        "6. Tras cada ciclo, regenerar la matriz de cobertura ejecutando el script.",
    ]
    base = 4 + len(info) + 2
    for i, line in enumerate(instructions):
        ws.cell(row=base + i, column=1, value="").alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=base + i, start_column=1, end_row=base + i, end_column=2)
        c = ws.cell(row=base + i, column=1)
        c.value = line
        c.font = cell_font()
        c.alignment = Alignment(wrap_text=True)


def build_plan(ws) -> None:
    headers = [label for _, label, _ in COLUMNS]

    for col_idx, (_key, _label, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30

    for r_idx, t in enumerate(TESTS, start=2):
        for c_idx, (key, _label, _w) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=t.get(key, ""))
            cell.font = cell_font()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW)

        ws.row_dimensions[r_idx].height = 60

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(TESTS) + 1}"


def build_cobertura(ws) -> None:
    ws.title = "Matriz cobertura"
    modulos = sorted({t["modulo"] for t in TESTS})
    tipos = sorted({t["tipo"] for t in TESTS})

    ws.column_dimensions["A"].width = 22
    for j, _t in enumerate(tipos, start=2):
        ws.column_dimensions[get_column_letter(j)].width = 15
    ws.column_dimensions[get_column_letter(len(tipos) + 2)].width = 12

    headers = ["Modulo \\ Tipo"] + tipos + ["Total"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.border = BORDER
    ws.row_dimensions[1].height = 26

    counter: Counter = Counter((t["modulo"], t["tipo"]) for t in TESTS)

    for i, m in enumerate(modulos, start=2):
        row_total = 0
        ws.cell(row=i, column=1, value=m).font = Font(bold=True, color=PRIMARY)
        ws.cell(row=i, column=1).alignment = Alignment(vertical="center")
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)

        for j, t in enumerate(tipos, start=2):
            n = counter.get((m, t), 0)
            row_total += n
            cell = ws.cell(row=i, column=j, value=n if n else "")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = cell_font()
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW)

        cell = ws.cell(row=i, column=len(tipos) + 2, value=row_total)
        cell.font = Font(bold=True, color=DARK_TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    last_row = len(modulos) + 2
    ws.cell(row=last_row, column=1, value="Total").font = Font(bold=True, color=HEADER_TEXT)
    ws.cell(row=last_row, column=1).fill = PatternFill("solid", fgColor=PRIMARY)
    ws.cell(row=last_row, column=1).border = BORDER
    grand_total = 0
    for j, t in enumerate(tipos, start=2):
        col_total = sum(1 for x in TESTS if x["tipo"] == t)
        grand_total += col_total
        c = ws.cell(row=last_row, column=j, value=col_total)
        c.font = Font(bold=True, color=HEADER_TEXT)
        c.fill = PatternFill("solid", fgColor=PRIMARY)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    c = ws.cell(row=last_row, column=len(tipos) + 2, value=grand_total)
    c.font = Font(bold=True, color=HEADER_TEXT)
    c.fill = PatternFill("solid", fgColor=PRIMARY)
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_resumen(wb.active)
    plan_ws = wb.create_sheet("Plan de pruebas")
    build_plan(plan_ws)
    cov_ws = wb.create_sheet("Matriz cobertura")
    build_cobertura(cov_ws)

    try:
        wb.save(OUTPUT)
        print(f"Excel generado: {OUTPUT} ({len(TESTS)} casos)")
    except PermissionError:
        fallback = OUTPUT.with_name(OUTPUT.stem + ".new" + OUTPUT.suffix)
        wb.save(fallback)
        print(
            "AVISO: el archivo original esta abierto en Excel.\n"
            f"Se guardo una copia nueva en: {fallback}\n"
            "Cierra Excel y vuelve a ejecutar para sobreescribir."
        )


if __name__ == "__main__":
    main()
