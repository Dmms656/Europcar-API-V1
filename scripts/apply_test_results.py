"""
Aplica los resultados de Contexto/resultados_pruebas.json sobre la hoja
"Plan de pruebas" del libro Plan_de_Pruebas_EUROPCAR_V1.xlsx.

Para cada caso ejecutado por scripts/run_tests.py escribe:
  - Estado (Pasada / Fallida / Bloqueada / No automatizado).
  - Responsable: "QA Automation".
  - Fecha ejecucion: yyyy-mm-dd.
  - Comentarios / evidencia: HTTP status, duracion en ms, observacion.

Tambien colorea la fila completa segun el estado para inspeccion visual.

Si el archivo .xlsx esta abierto en Excel y no se puede sobrescribir, se
guarda como Plan_de_Pruebas_EUROPCAR_V1.evidencias.xlsx en la misma carpeta.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

REPO_ROOT = Path(__file__).resolve().parent.parent
RESULTS = REPO_ROOT / "Contexto" / "resultados_pruebas.json"
PLAN = REPO_ROOT / "Contexto" / "Plan_de_Pruebas_EUROPCAR_V1.xlsx"
PLAN_OUT = REPO_ROOT / "Contexto" / "Plan_de_Pruebas_EUROPCAR_V1.xlsx"
PLAN_FALLBACK = REPO_ROOT / "Contexto" / "Plan_de_Pruebas_EUROPCAR_V1.evidencias.xlsx"

SHEET = "Plan de pruebas"
RESUMEN_SHEET = "Resumen"

ESTADO_COLORS = {
    "Pasada":           "DCFCE7",
    "Fallida":          "FEE2E2",
    "Bloqueada":        "FFEDD5",
    "No automatizado":  "E5E7EB",
    "Pendiente":        "FFFFFF",
}
ESTADO_TEXT = {
    "Pasada":           "166534",
    "Fallida":          "991B1B",
    "Bloqueada":        "9A3412",
    "No automatizado":  "374151",
    "Pendiente":        "1B2631",
}


def column_index_by_header(ws, headers: list[str]) -> dict[str, int]:
    out: dict[str, int] = {}
    target = {h.strip().lower(): h for h in headers}
    for col_idx, cell in enumerate(ws[1], start=1):
        v = (cell.value or "").strip().lower() if isinstance(cell.value, str) else ""
        if v in target:
            out[target[v]] = col_idx
    return out


def build_evidence(item: dict) -> str:
    bits: list[str] = []
    http = item.get("http_status")
    if http != "" and http is not None:
        bits.append(f"HTTP {http}")
    if item.get("duration_ms"):
        bits.append(f"{item['duration_ms']} ms")
    bits.append(item.get("comentario") or "")
    return " | ".join(b for b in bits if b)


def apply_results() -> Path:
    if not RESULTS.exists():
        raise SystemExit(f"No existe {RESULTS}; ejecuta primero scripts/run_tests.py")

    data = json.loads(RESULTS.read_text(encoding="utf-8"))
    by_id = {row["id"]: row for row in data}

    wb = load_workbook(PLAN)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Falta la hoja '{SHEET}' en el libro {PLAN}")

    ws = wb[SHEET]
    needed = ["ID", "Estado", "Responsable", "Fecha ejecucion",
              "Comentarios / evidencia", "Severidad si falla"]
    cols = column_index_by_header(ws, needed)
    missing = [n for n in needed if n not in cols]
    if missing:
        raise SystemExit(f"Faltan columnas en la hoja '{SHEET}': {missing}")

    counters: dict[str, int] = {}
    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        tid = ws.cell(row=row_idx, column=cols["ID"]).value
        if not tid:
            continue
        item = by_id.get(str(tid).strip())
        if not item:
            continue
        estado = item["status"]
        counters[estado] = counters.get(estado, 0) + 1
        evidence = build_evidence(item)

        ws.cell(row=row_idx, column=cols["Estado"]).value = estado
        ws.cell(row=row_idx, column=cols["Responsable"]).value = "QA Automation"
        ws.cell(row=row_idx, column=cols["Fecha ejecucion"]).value = item["fecha"]
        ws.cell(row=row_idx, column=cols["Comentarios / evidencia"]).value = evidence

        if estado == "Pasada":
            ws.cell(row=row_idx, column=cols["Severidad si falla"]).value = "-"

        fill = PatternFill("solid", fgColor=ESTADO_COLORS.get(estado, "FFFFFF"))
        font_color = ESTADO_TEXT.get(estado, "1B2631")
        for c_idx in (cols["Estado"], cols["Comentarios / evidencia"]):
            ws.cell(row=row_idx, column=c_idx).fill = fill
            ws.cell(row=row_idx, column=c_idx).font = Font(
                name="Calibri", size=10, bold=(c_idx == cols["Estado"]),
                color=font_color)
        updated += 1

    # Resumen ejecutivo en la primera hoja
    if RESUMEN_SHEET in wb.sheetnames:
        rs = wb[RESUMEN_SHEET]
        anchor_row = (rs.max_row or 1) + 2
        rs.cell(row=anchor_row, column=2).value = "Resultados de la ejecucion"
        rs.cell(row=anchor_row, column=2).font = Font(
            name="Calibri", size=14, bold=True, color="0F4C81")
        rs.cell(row=anchor_row + 1, column=2).value = (
            f"Casos actualizados: {updated} de {len(by_id)}")
        order = ["Pasada", "Fallida", "Bloqueada", "No automatizado"]
        for i, k in enumerate(order, start=anchor_row + 2):
            rs.cell(row=i, column=2).value = k
            rs.cell(row=i, column=3).value = counters.get(k, 0)
            rs.cell(row=i, column=2).fill = PatternFill(
                "solid", fgColor=ESTADO_COLORS.get(k, "FFFFFF"))
            rs.cell(row=i, column=2).font = Font(
                name="Calibri", size=11, bold=True,
                color=ESTADO_TEXT.get(k, "1B2631"))

    try:
        wb.save(PLAN_OUT)
        out = PLAN_OUT
    except PermissionError:
        wb.save(PLAN_FALLBACK)
        out = PLAN_FALLBACK
        print(f"[aviso] El archivo original esta abierto. Se guardo en {out.name}.")

    print(f"Casos con evidencia: {updated} de {len(by_id)}")
    for k in ("Pasada", "Fallida", "Bloqueada", "No automatizado"):
        print(f"  {k:18s} {counters.get(k, 0)}")
    print(f"Libro generado: {out}")
    return out


if __name__ == "__main__":
    apply_results()
